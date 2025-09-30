from openpyxl import load_workbook
from awswrangler import redshift
from io import BytesIO
import pandas as pd
import numpy as np
import os
from utils import get_session, get_secret, get_rs_conn, send_slack_notification, get_webhook, get_logger

# -------------------------------------- Functions --------------------------------------

# Function to initialize global varaibles
def init_globals():
    global logger, session, bucket_name, s3, creds, conn, session
    logger = get_logger()
    creds = {'db': 'wbx_data', 'cluster_id': 'wbx-data', 'region': 'us-west-2'}
    bucket_name = 'wbx-data.redshift-unload'

    # Initializing Botocore session
    logger.info("Initializing Botocore session ...")
    session = get_session(creds)
    s3 = session.resource('s3')

    # Establishing Redshift connection
    logger.info("Establishing Redshift connection ...")
    rs_secret = get_secret(os.environ['RS_SECRET'], session, creds['region'])
    conn = get_rs_conn(rs_secret)


# Function to load manual transactions
def load_manual_transactions(table_name):
    global logger, bucket_name, s3, creds, conn
    
    # Get all files in the bucket
    bucket = s3.Bucket(bucket_name)
    prefix = f'raw/{table_name}/'
    files = [obj.key for obj in bucket.objects.filter(Prefix=prefix) if obj.key.endswith('.xlsx')]

    all_dfs = []
    # Read all files and append to a list
    logger.info("Fetching data from S3 ...")
    for file in files:
        obj = s3.Object(bucket_name, file)
        body = obj.get()['Body'].read()

        # Load the workbook from the in-memory file object
        file_stream = BytesIO(body)
        workbook = load_workbook(file_stream)
        
        dfs = []
        # Iterate over each sheet
        for s in workbook.sheetnames:
            sheet = workbook[s]
            data = sheet.values
            df = pd.DataFrame(data)
            df.columns = df.iloc[0]
            df = df[1:]
            df.reset_index(drop=True, inplace=True)
            df = df.dropna(how='all')
            df['Date'] = pd.to_datetime(df['Date']).dt.date
            df['Account ID'] = df['Account ID'].apply(lambda x: f"{int(x):.0f}" if not pd.isna(x) else np.nan)
            df = df[['Date', 'Account ID', 'Product', 'Amount', 'Source - Tableau (additional item)']]
            dfs.append(df)

        # Concatenate all DataFrames into one
        combined_df = pd.concat(dfs, ignore_index=True)
        all_dfs.append(combined_df)

    # Combine all dataframes
    all_combined_df = pd.concat(all_dfs, ignore_index=True)
    all_combined_df.columns = ['date_created', 'account_id', 'product', 'total', 'source']
    all_combined_df.dropna(subset=['account_id'], inplace= True)
    all_combined_df['account_id'] = all_combined_df['account_id'].astype(int)
  
    # Drop table
    conn.cursor().execute(f"DROP TABLE IF EXISTS {creds['db']}.{table_name} CASCADE;")

    # Write to Redshift
    logger.info("Writing to Redshift ...")
    redshift.copy(
                df = all_combined_df,
                path = f's3://{bucket_name}/raw/temp.csv',
                con = conn,
                table = table_name,
                schema = creds['db'],
                mode = "overwrite",
                overwrite_method = "drop",
                boto3_session = session
            )
    
    
# -------------------------------------- Start --------------------------------------

def main():
    init_globals()
    logger.info("Starting ...")

    webhook_url = get_webhook(session, creds)
    send_slack_notification(webhook_url, 'Loading manual transactions :loading_dots:')

    table_names = ['manual_transactions', 'manual_transactions_pp']

    for t in table_names:
        load_manual_transactions(t)

    # Grant permissions to the new tables
    conn.cursor().execute(f"GRANT ALL ON ALL TABLES in SCHEMA {creds['db']} to group admin;")
    conn.commit()
    conn.close()
    
    logger.info("Done!")
    send_slack_notification(webhook_url, 'Manual transactions loaded :white_check_mark:')


if __name__ == '__main__':
    main()
